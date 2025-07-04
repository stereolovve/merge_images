# Fichas.py
# Script para inserir os dados do banco de dados na planilha de fichas de acessos

# Leitura do banco de dados
import pandas as pd
import shutil
from openpyxl import load_workbook
from openpyxl.styles import Font
import os

# Ler o banco de dados
df = pd.read_excel(r"C:\Users\lucas.melo\merge_images\base\Lista_050.xlsx")

# Verificar IDs duplicados
ids_duplicados = df[df['ID'].duplicated(keep=False)]['ID'].unique()
if len(ids_duplicados) > 0:
    print(f"Encontrados {len(ids_duplicados)} IDs duplicados: {ids_duplicados}")

# Contador de arquivos criados
arquivos_criados = 0
erros = []

# Para cada ponto, gerar uma ficha de acesso
for i, row in df.iterrows():
    try:
        # Guardar o identificador para colocar como nome da ficha
        identificacao = row["ID"]

        # Verifica se o ID é duplicado
        if identificacao in ids_duplicados:
            # Se for duplicado, inclui o PONTO no nome do arquivo
            ponto = str(row["PONTO"]).replace("/", "-")  # Substitui / por - para evitar problemas com nomes de arquivo
            new_ficha = f"fichas/{identificacao}_{ponto}.xlsx"
        else:
            # Se não for duplicado, mantém apenas o ID
            new_ficha = f"fichas/{identificacao}.xlsx"

        # Copia o modelo de ficha
        shutil.copy("Ficha_050.xlsx", new_ficha)

        wb = load_workbook(new_ficha)
        ws = wb.active

        # Criar fonte preta
        black_font = Font(color="000000")

        # Preenche os campos da ficha
        ws["Y4"]  = row["PONTO"]
        ws["Y4"].font = black_font
        ws["W3"]  = row["ID"]
        ws["W3"].font = black_font
        ws["E7"]  = row["RODOVIA"]
        ws["E7"].font = black_font
        ws["N7"]  = row["KM"]
        ws["N7"].font = black_font
        ws["V7"]  = row["SENTIDO"]
        ws["V7"].font = black_font
        ws["E8"]  = row["MUNICÍPIO"]
        ws["E8"].font = black_font
        ws["V8"]  = row["UF"]
        ws["V8"].font = black_font
        ws["I9"]  = row["LATITUDE"]
        ws["I9"].font = black_font
        ws["U9"]  = row["LONGITUDE"]
        ws["U9"].font = black_font
        ws["E12"] = row["REGULARIZAÇÃO"]
        ws["E12"].font = black_font
        ws["E13"] = row["ACIDENTES"]
        ws["E13"].font = black_font
        ws["S13"] = row["AVALIAÇÃO ETAPA 1"]
        ws["S13"].font = black_font
        ws["H16"] = row["VGD"]
        ws["H16"].font = black_font
        ws["Q16"] = row["VEÍCULO TIPO"]
        ws["Q16"].font = black_font
        ws["X16"] = row["TIPIFICAÇÃO (IPR728)"]
        ws["X16"].font = black_font
        ws["H17"] = row["DISPOSITIVOS OPERACIONAIS"]
        ws["H17"].font = black_font
        ws["Q17"] = row["CURVA HORIZONTAL"]
        ws["Q17"].font = black_font
        ws["X17"] = row["CURVA VERTICAL"]
        ws["X17"].font = black_font
        ws["E18"] = row["ACIDENTES (3 ANOS)"]
        ws["E18"].font = black_font
        ws["S18"] = row["AVALIAÇÃO ETAPA 2"]
        ws["S18"].font = black_font

        wb.save(new_ficha)
        arquivos_criados += 1
        print(f"Criado arquivo {new_ficha}")

    except Exception as e:
        erros.append(f"Erro na linha {i+2}: {str(e)}")
        print(f"Erro na linha {i+2}: {str(e)}")

# Relatório final
print("\nRelatório final:")
print(f"Total de linhas no arquivo: {len(df)}")
print(f"Total de arquivos criados: {arquivos_criados}")
print(f"Total de erros: {len(erros)}")

if erros:
    print("\nDetalhes dos erros:")
    for erro in erros:
        print(erro)


